'''
面向多条切割线的XXXX作业问题

在1.0的基础上变化了local search的方法
'''
import math
import os.path
import time
import re
import numpy as np
import xlrd
import gc
import objgraph
from itertools import product
from copy import deepcopy
from random import choice
from RSP import RSP, simplifiedRSP
from OriginalModel import originalModel
from openpyxl import Workbook
from openpyxl import load_workbook
'''
基本参数
'''
S = int()#堆位数量
N = int()#任务数量，出库钢板数量，出库阶段数量
L = int()#切割线数量
H = int()#堆位最大高度限制
DT = float()#从堆场移动到切割线的时间
CT = dict()#切割工时数据
ES = dict()#每个堆位空板位数量
PlatesInStack = dict()#每个堆位内的出库钢板编号，从前往后从下到上
B = dict()#每个出库钢板上方阻挡板数量
MOT = dict()#每个出库钢板出库时间的下界
RT = np.empty((S, S), dtype=float)#两个堆位翻板时间
OT = dict()#出库钢板的提取时间
P = np.zeros((N, S), dtype=float)#01量说明任务n在堆位s
CutLineOfPlate = dict()#每个任务所在切割线的编号，从0开始
StackOfPlate = dict()#每个任务所在的堆位编号，从0开始
K = np.zeros((N, N), dtype=float)#判断两个任务是否在同一个切割线
O = dict()#每个任务上方出库钢板编号

def originalData(url):
    global S, N, L, DT, CT, PlatesInStack, B, MOT, RT, OT, P, CutLineOfPlate, StackOfPlate, H, K, O, ES
    workbook = xlrd.open_workbook(url)
    basicDataSheet = workbook.sheet_by_index(0)
    '''
    基础参数
    '''
    S = int(basicDataSheet.cell(0, 1).value)
    N = int(basicDataSheet.cell(1, 1).value)
    H = int(basicDataSheet.cell(2, 1).value)
    L = int(basicDataSheet.cell(3, 1).value)
    rowNumber = int(basicDataSheet.cell(4, 1).value)  # 堆场多少行
    colNumber = int(basicDataSheet.cell(5, 1).value)  # 堆场多少列
    timeFor1Row = float(basicDataSheet.cell(6, 1).value)  # 移动一行的时间
    timeFor1Col = float(basicDataSheet.cell(7, 1).value)  # 移动一列的时间
    eta = float(basicDataSheet.cell(8, 1).value)  # 一次抬升的时间
    DT = float(basicDataSheet.cell(9, 1).value)
    CT = dict()
    PlatesInStack = dict()  # 每个堆位内的出库钢板编号，从前往后从下到上
    B = dict()  # 每个出库钢板上方阻挡板数量
    MOT = dict()  # 每个出库钢板出库时间的下界
    RT = np.empty((S, S), dtype=float)  # 两个堆位翻板时间
    OT = dict()  # 出库钢板的提取时间
    P = np.zeros((N, S), dtype=float)  # 01量说明任务n在堆位s
    CutLineOfPlate = dict()  # 每个任务所在切割线的编号，从0开始
    StackOfPlate = dict()  # 每个任务所在的堆位编号，从0开始
    K = np.zeros((N, N), dtype=float)  # 判断两个任务是否在同一个切割线
    O = dict()  # 每个任务上方出库钢板编号
    ES = dict()  # 每个堆位空板位数量

    '''
    根据基础参数得到一些东西，包括：
    RT：两个堆位之间的移动时间
    ETForStack:堆位和出口的移动时间（用于OT_i）
    '''
    RT = np.empty((S, S), dtype=float)
    ETForStack = dict()
    row_e = rowNumber - 1
    col_e = colNumber
    for i, j in product(range(S), range(S)):
        row_i = math.floor(i / colNumber)
        row_j = math.floor(j / colNumber)
        col_i = i - row_i * colNumber
        col_j = j - row_j * colNumber
        RT[i, j] = 0 if i == j else abs(row_i - row_j) * timeFor1Row + abs(col_i - col_j) * timeFor1Col + eta
        ETForStack[i] = abs(row_i - row_e) * timeFor1Row + abs(col_i - col_e) * timeFor1Col + eta
    '''
    切割工时数据:CT
    每个任务所在的切割线：CutLineOFPlate
    两个任务是否在同一条切割线：K
    '''
    cuttingTimeSheet = workbook.sheet_by_index(1)
    a = N / L
    K = np.zeros((N, N), dtype=float)
    for job in range(N):
        CT[job] = float(cuttingTimeSheet.cell(job + 1, 1).value)
        CutLineOfPlate[job] = math.floor(job/a)
    for i,j in product(range(N), range(N)):
        if CutLineOfPlate[i] == CutLineOfPlate[j]: K[i,j] = 1
    '''
    存储信息，得到：
    ES：每个堆位的初始空板位数量
    PlatesInStack:每个堆位出库钢板编号，从下到上排序
    B:每个出库钢板上方的阻挡板数量
    MOT:每个出库时间的下界
    P:堆位是否有板i的01量
    StackOfPlate: 任务所在堆位位置
    O:每个任务上方的钢板
    '''
    P = np.zeros((N, S), dtype=float)
    storageSheet = workbook.sheet_by_index(2)
    for s in range(S):
        #先拿数据
        stackConfig = []
        for h in range(len(storageSheet.row(s))):
            if storageSheet.cell(s, h).value != '':stackConfig.append(int(storageSheet.cell(s, h).value))
        #空板位数量
        ES[s] = H - len(stackConfig)
        #B, PlatesInStack; bTop为临时数据
        PlatesInStack[s] = []
        bTop = dict()
        for plate in stackConfig:
            if plate != -1:
                PlatesInStack[s].append(plate)
                StackOfPlate[plate] = s
                P[plate, s] = 1
                B[plate] = 0
                bTop[plate] = 0
            elif len(PlatesInStack[s]) > 0:
                for plate in PlatesInStack[s]:B[plate] = B[plate] + 1
                bTop[PlatesInStack[s][-1]] = bTop[PlatesInStack[s][-1]] + 1
        #确定O
        tempPlateInStackS = deepcopy(PlatesInStack[s])
        while len(tempPlateInStackS) > 1:
            plate = tempPlateInStackS.pop(0)
            O[plate] = []
            for otherPlate in tempPlateInStackS: O[plate].append(otherPlate)
        del tempPlateInStackS

        #OT,MOT值
        #首先确定和当前堆位翻板最近的时间
        minRelTime = 1000
        for time in RT[s]:
            if time != 0.0 and time < minRelTime: minRelTime = time
        for plate in PlatesInStack[s]:
            OT[plate] = ETForStack[s]
            MOT[plate] = bTop[plate] * minRelTime + OT[plate]

if __name__ == '__main__':
    re_pattern = 'b_3'
    file_name = 'origin model result b3.xlsx'

    dataPath = './data/experiment 1/'
    if os.path.exists(dataPath):
        files = os.listdir(dataPath)

    for eachfile in files:
        if re.search(re_pattern, eachfile)!= None:
            print('\n\n***********************************\n', eachfile, '\n________________________________')
            originalData(dataPath + eachfile)
            # lb
            lowPlate = []
            for s in range(S):
                if len(PlatesInStack[s]) > 0: lowPlate.append(PlatesInStack[s][0])
            lb1 = sum([MOT[i] for i in MOT.keys()]) + min([CT[i] for i in lowPlate])


            CTPerLine = [0 for line in range(L)]
            MinOTPerLine = [10000 for line in range(L)]
            for job in range(N):
                CTPerLine[CutLineOfPlate[job]] += CT[job]
                if MOT[job] < MinOTPerLine[CutLineOfPlate[job]]: MinOTPerLine[CutLineOfPlate[job]] = MOT[job]
            lb2 = max([CTPerLine[l] + MinOTPerLine[l] + DT for l in range(L)])
            lb = max(lb1, lb2)
            print('lb1=',lb1,'lb2=',lb2,'\n________________________________')


            #original Model
            while True:
                res, status, bound, time = originalModel(S, N, O, H, P, B, ES, OT, RT, L, DT, CT, K, lb, 1800)
                print(res)
                print(status)
                print(bound)
                # write results into excel
                wb = load_workbook(dataPath + file_name)
                sheet = wb.active
                nrow = sheet.max_row + 1
                sheet.cell(row=nrow, column=1, value=eachfile)
                sheet.cell(row=nrow, column=2, value=res)
                sheet.cell(row=nrow, column=3, value=status)
                sheet.cell(row=nrow, column=4, value=bound)
                sheet.cell(row=nrow, column=5, value=time)
                os.remove(dataPath + file_name)
                wb.save(dataPath + file_name)

                #del S, N, O, H, P, B, ES, OT, RT, L, DT, CT, K, lb
                break

