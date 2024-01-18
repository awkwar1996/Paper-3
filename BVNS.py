'''
面向多条切割线的XXXX作业问题

20230210在1.0的基础上变化了local search的方法
20230228在2.0的基础上增加计算各条线切割总工时的标准差
'''
import math
import os.path
import time
import re
import numpy as np
import xlrd
import gc
import objgraph
from itertools import product
from copy import deepcopy
from random import choice
from RSP import RSP, simplifiedRSP
#from OriginalModel import originalModel
from openpyxl import Workbook
from openpyxl import load_workbook
'''
基本参数
'''
S = int()#堆位数量
N = int()#任务数量，出库钢板数量，出库阶段数量
L = int()#切割线数量
H = int()#堆位最大高度限制
DT = float()#从堆场移动到切割线的时间
CT = dict()#切割工时数据
ES = dict()#每个堆位空板位数量
PlatesInStack = dict()#每个堆位内的出库钢板编号，从前往后从下到上
B = dict()#每个出库钢板上方阻挡板数量
MOT = dict()#每个出库钢板出库时间的下界
RT = np.empty((S, S), dtype=float)#两个堆位翻板时间
OT = dict()#出库钢板的提取时间
P = np.zeros((N, S), dtype=float)#01量说明任务n在堆位s
CutLineOfPlate = dict()#每个任务所在切割线的编号，从0开始
StackOfPlate = dict()#每个任务所在的堆位编号，从0开始
K = np.zeros((N, N), dtype=float)#判断两个任务是否在同一个切割线
O = dict()#每个任务上方出库钢板编号

def originalData(url):
    global S, N, L, DT, CT, PlatesInStack, B, MOT, RT, OT, P, CutLineOfPlate, StackOfPlate, H, K, O, ES
    workbook = xlrd.open_workbook(url)
    basicDataSheet = workbook.sheet_by_index(0)
    '''
    基础参数
    '''
    S = int(basicDataSheet.cell(0, 1).value)
    N = int(basicDataSheet.cell(1, 1).value)
    H = int(basicDataSheet.cell(2, 1).value)
    L = int(basicDataSheet.cell(3, 1).value)
    rowNumber = int(basicDataSheet.cell(4, 1).value)  # 堆场多少行
    colNumber = int(basicDataSheet.cell(5, 1).value)  # 堆场多少列
    timeFor1Row = float(basicDataSheet.cell(6, 1).value)  # 移动一行的时间
    timeFor1Col = float(basicDataSheet.cell(7, 1).value)  # 移动一列的时间
    eta = float(basicDataSheet.cell(8, 1).value)  # 一次抬升的时间
    DT = float(basicDataSheet.cell(9, 1).value)
    CT = dict()
    PlatesInStack = dict()  # 每个堆位内的出库钢板编号，从前往后从下到上
    B = dict()  # 每个出库钢板上方阻挡板数量
    MOT = dict()  # 每个出库钢板出库时间的下界
    RT = np.empty((S, S), dtype=float)  # 两个堆位翻板时间
    OT = dict()  # 出库钢板的提取时间
    P = np.zeros((N, S), dtype=float)  # 01量说明任务n在堆位s
    CutLineOfPlate = dict()  # 每个任务所在切割线的编号，从0开始
    StackOfPlate = dict()  # 每个任务所在的堆位编号，从0开始
    K = np.zeros((N, N), dtype=float)  # 判断两个任务是否在同一个切割线
    O = dict()  # 每个任务上方出库钢板编号
    ES = dict()  # 每个堆位空板位数量

    '''
    根据基础参数得到一些东西，包括：
    RT：两个堆位之间的移动时间
    ETForStack:堆位和出口的移动时间（用于OT_i）
    '''
    RT = np.empty((S, S), dtype=float)
    ETForStack = dict()
    row_e = rowNumber - 1
    col_e = colNumber
    for i, j in product(range(S), range(S)):
        row_i = math.floor(i / colNumber)
        row_j = math.floor(j / colNumber)
        col_i = i - row_i * colNumber
        col_j = j - row_j * colNumber
        RT[i, j] = 0 if i == j else abs(row_i - row_j) * timeFor1Row + abs(col_i - col_j) * timeFor1Col + eta
        ETForStack[i] = abs(row_i - row_e) * timeFor1Row + abs(col_i - col_e) * timeFor1Col + eta
    '''
    切割工时数据:CT
    每个任务所在的切割线：CutLineOFPlate
    两个任务是否在同一条切割线：K
    '''
    cuttingTimeSheet = workbook.sheet_by_index(1)

    K = np.zeros((N, N), dtype=float)
    for job in range(N):
        CT[job] = float(cuttingTimeSheet.cell(job + 1, 1).value)
        CutLineOfPlate[job] = int(cuttingTimeSheet.cell(job + 1, 2).value)
    for i,j in product(range(N), range(N)):
        if CutLineOfPlate[i] == CutLineOfPlate[j]: K[i,j] = 1
    '''
    存储信息，得到：
    ES：每个堆位的初始空板位数量
    PlatesInStack:每个堆位出库钢板编号，从下到上排序
    B:每个出库钢板上方的阻挡板数量
    MOT:每个出库时间的下界
    P:堆位是否有板i的01量
    StackOfPlate: 任务所在堆位位置
    O:每个任务上方的钢板
    '''
    P = np.zeros((N, S), dtype=float)
    storageSheet = workbook.sheet_by_index(2)
    for s in range(S):
        #先拿数据
        stackConfig = []
        for h in range(len(storageSheet.row(s))):
            if storageSheet.cell(s, h).value != '':stackConfig.append(int(storageSheet.cell(s, h).value))
        #空板位数量
        ES[s] = H - len(stackConfig)
        #B, PlatesInStack; bTop为临时数据
        PlatesInStack[s] = []
        bTop = dict()
        for plate in stackConfig:
            if plate != -1:
                PlatesInStack[s].append(plate)
                StackOfPlate[plate] = s
                P[plate, s] = 1
                B[plate] = 0
                bTop[plate] = 0
            elif len(PlatesInStack[s]) > 0:
                for plate in PlatesInStack[s]:B[plate] = B[plate] + 1
                bTop[PlatesInStack[s][-1]] = bTop[PlatesInStack[s][-1]] + 1
        #确定O
        tempPlateInStackS = deepcopy(PlatesInStack[s])
        while len(tempPlateInStackS) > 1:
            plate = tempPlateInStackS.pop(0)
            O[plate] = []
            for otherPlate in tempPlateInStackS: O[plate].append(otherPlate)
        del tempPlateInStackS

        #OT,MOT值
        #首先确定和当前堆位翻板最近的时间
        minRelTime = 1000
        for time in RT[s]:
            if time != 0.0 and time < minRelTime: minRelTime = time
        for plate in PlatesInStack[s]:
            OT[plate] = ETForStack[s]
            MOT[plate] = bTop[plate] * minRelTime + OT[plate]
def stackBasedInitialSolution():
    pi = []
    stackCandidate = []
    #排除没有任务钢板的堆位
    for s in range(S):
        if len(PlatesInStack[s]) > 0: stackCandidate.append(s)
    #循环添加堆位
    while len(pi) < N:
        if len(pi) == 0:#一开始，选择BP最少的堆位
            minB = 100
            chosenStack = -1
            for s in stackCandidate:
                sB = max([B[plate] for plate in PlatesInStack[s]])
                if sB < minB:
                    chosenStack = s
                    minB = sB
        else:#选择和之前一个堆位最近的堆位，多个的话选择BP数量最少的。
            formerChosenStack = chosenStack
            minRT = 1000
            for s in stackCandidate:
                #选择距离最近的
                if RT[s, formerChosenStack] < minRT:
                    chosenStack = s
                    minRT = RT[chosenStack, formerChosenStack]
                #选择BP数量最少的
                elif RT[s, formerChosenStack] == minRT and max([B[plate] for plate in PlatesInStack[s]]) < max(
                    [B[plate] for plate in PlatesInStack[chosenStack]]):
                    chosenStack = s
        pi.extend(PlatesInStack[chosenStack][::-1])
        stackCandidate.remove(chosenStack)
    return pi
def noDelayInitialSolution():
    '''
    首先tL将切割时间差先放上去
    每一轮每条线只选择一个MOT最小的任务
    在L条线的任务中，选择delay的任务
    多个delay任务，计算当前任务出库完成时间+DT+后面任务的切割时间和，选择最大的任务
    '''
    #初始化
    pi = []
    tY = 0
    tL = [0 for l in range(L)]
    tempPlatesInStack = deepcopy(PlatesInStack)
    #计算每条线剩余切割时间
    remainedCutTime = [0 for l in range(L)]
    for n in range(N):
        remainedCutTime[CutLineOfPlate[n]] += CT[n]
    #循环确定任务
    while len(pi) < N:
        candiPlateForLine = [-1 for l in range(L)]
        chosenPlate = -1
        #每个堆位循环，确定每条线的candiPlateForLine
        for s in range(S):
            #当前堆位没有钢板
            if len(tempPlatesInStack[s]) == 0: continue
            plate = tempPlatesInStack[s][-1]
            line = CutLineOfPlate[plate]
            #先确定plate是否为当前切割线最快出库的,不是的话剔除
            if candiPlateForLine[line] != -1 and MOT[plate] > MOT[candiPlateForLine[line]]: continue
            #把plate作为candiPlateForLine
            candiPlateForLine[line] = plate
        # 确定chosenPlate
        for l in range(L):
            if candiPlateForLine[l] == -1:
                continue
            plate = candiPlateForLine[l]
            if chosenPlate == -1:
                chosenPlate = plate
                continue
            # 1.当前任务紧急和备选任务不紧急
            if tY + MOT[plate] + DT >= tL[l] and tY + MOT[chosenPlate] + DT < tL[CutLineOfPlate[chosenPlate]]:
                chosenPlate = plate
            # 3.都是紧急任务
            elif tY + MOT[plate] + DT >= tL[l]: #and tY + MOT[chosenPlate] + DT >= tL[CutLineOfPlate[chosenPlate]]:
                # 都是紧急,以到达时间计算
                if MOT[plate] + remainedCutTime[l] > MOT[chosenPlate] + remainedCutTime[CutLineOfPlate[chosenPlate]]:
                    chosenPlate = plate
            # 4.都是不紧急任务
            elif tY + MOT[plate] + DT < tL[l] and tY + MOT[chosenPlate] + DT < tL[CutLineOfPlate[chosenPlate]]:
                # 都不紧急，以切割线时间计算
                if tL[l] + remainedCutTime[l] > tL[CutLineOfPlate[chosenPlate]] + remainedCutTime[
                    CutLineOfPlate[chosenPlate]]:
                    chosenPlate = plate
        #为了方便，建立chosenLine
        chosenLine = CutLineOfPlate[chosenPlate]
        #加上选择
        pi.append(chosenPlate)
        #从堆位中删除
        tempPlatesInStack[StackOfPlate[chosenPlate]].remove(chosenPlate)
        #更新时间
        tY += MOT[chosenPlate]
        tL[chosenLine] = max(tL[chosenLine], tY + DT) + CT[chosenPlate]
        #更新remainedCutTime
        remainedCutTime[chosenLine] -= CT[chosenPlate]
    return pi
def randomInitialSolution():
    pi = []
    tempPlatesIStack = deepcopy(PlatesInStack)
    while len(pi) < N:
        candidateSet = []
        for s in range(S):
            if len(tempPlatesIStack[s]) > 0: candidateSet.append(tempPlatesIStack[s][-1])
        chosenPlate = choice(candidateSet)
        pi.append(chosenPlate)
        tempPlatesIStack[StackOfPlate[chosenPlate]].remove(chosenPlate)
    return pi
#输出pi和pi的结果
def localSearch(pi):
    #确定bottleneck上的delay plate合集，要求不能在plateTabu中, 第一个出库任务在里头
    def getDelayPlate(pi, outboundTime, cutTime, plateTabu):
        #1.确定bottleneck line
        cLMax = [0 for l in range(L)]
        for plate in pi[::-1]:
            if cLMax[CutLineOfPlate[plate]] == 0:cLMax[CutLineOfPlate[plate]] = cutTime[plate]
            if 0 not in cLMax: break#这里假设所有线都有任务
        bottleneckLine = cLMax.index(max(cLMax))
        #2.找到bottleneck line上所有delayed plate，按照出库顺序依次排列
        #属于Tabu的不要
        bigOmega = []
        formerPlate = -1
        for plate in pi:
            if CutLineOfPlate[plate] != bottleneckLine: continue
            #在Tabu里头的排除
            if plate not in plateTabu:
                #第一个任务，直接加
                if formerPlate == -1: bigOmega.append(plate)
                #后续任务，计算是不是delay plate
                elif outboundTime[plate] + DT > cutTime[formerPlate]: bigOmega.append(plate)
            formerPlate = plate
        return bigOmega
    def changePosition(i, pi, outboundTime, cutTime, nTabu):
        global CutLineOfPlate, StackOfPlate
        def judge(pi, i, n, outboundTime, cutTime):
            #1. 确定该任务的预期移动增加基准值
            adding = outboundTime[i] - outboundTime[pi[pi.index(i) - 1]]

            #2. 计算每条线的makespan
            cLMax = [0 for l in range(L)]
            for plate in pi[::-1]:
                if cLMax[CutLineOfPlate[plate]] == 0: cLMax[CutLineOfPlate[plate]] = cutTime[plate]
                if 0 not in cLMax: break

            #3. 计算每条线的预期增加值(i之前的增加量)
            candidateL = dict()
            formerCutTimeL = dict()
            #3.1 先确定有几条线有任务
            for plate in pi[n:pi.index(i)]:
                if CutLineOfPlate[plate] != CutLineOfPlate[i]:
                    candidateL[CutLineOfPlate[plate]] = 0
                    formerCutTimeL[CutLineOfPlate[plate]] = 0
            #3.2 从前往后，计算预期增加量减少值，用递推公式计算
            for plate in pi[:pi.index(i)]:
                tempL = CutLineOfPlate[plate]
                if tempL in candidateL.keys():
                    if plate in pi[n:pi.index(i)]:
                        candidateL[tempL] = adding if outboundTime[plate] + DT - formerCutTimeL[tempL] >= 0 \
                            else max(candidateL[tempL], adding - formerCutTimeL[tempL] + outboundTime[plate] + DT)
                    formerCutTimeL[tempL] = cutTime[plate]
            #3.3去掉预期增加值小于gap差的线
            tempCandidateL = deepcopy(candidateL)
            for lIndex in tempCandidateL.keys():
                if candidateL[lIndex] < max(cLMax) - cLMax[lIndex]:
                    del candidateL[lIndex]
            del tempCandidateL
            if len(candidateL.keys()) == 0: return True

            #4. 计算每条线的结果
            for plate in pi[pi.index(i):]:
                tempL = CutLineOfPlate[plate]
                if tempL in candidateL.keys():
                    if outboundTime[plate] + DT > formerCutTimeL[tempL]:
                        candidateL[tempL] = max(0, candidateL[tempL] - outboundTime[plate] - DT + formerCutTimeL[tempL])
                    formerCutTimeL[tempL] = cutTime[plate]
            #5.输出结果
            for lIndex in candidateL.keys():
                if candidateL[lIndex] > max(cLMax) - cLMax[lIndex]: return False
            return True


        #函数分割线——————————
        #第一个任务，直接否定
        if pi.index(i) == 0: return False, -1
        n = -1
        j = pi.index(i)
        #确定n
        while j >= 1:
            if CutLineOfPlate[pi[j - 1]] == CutLineOfPlate[i] or StackOfPlate[pi[j - 1]] == StackOfPlate[i] or \
                    j - 1 in nTabu:
                break
            elif judge(pi, i, j - 1, outboundTime, cutTime):
                n = j - 1
            j -= 1
        #n不更新，否定,其他就确定
        return (False, n) if n == -1 else (True, n)
    def insertion(pi, i, n):
        indexI = pi.index(i)
        pi.insert(n, pi.pop(indexI))
        return pi
    #函数主体
    plateTabu = []
    bestRes, outboundTime, cutTime = RSP(pi, CutLineOfPlate, StackOfPlate, L, N, S, DT, CT, OT, ES, P, B, RT)
    bigOmega = getDelayPlate(pi, outboundTime, cutTime, plateTabu)
    while len(bigOmega) != 0:
        nTabu = []
        #i是delay plate
        i = bigOmega[0]
        #调整位置
        while True:
            #确定更换的位置，如果hasPositionFlag = False, 说明没有找到合适的位置。
            hasPositionFlag, n = changePosition(i, pi, outboundTime, cutTime, nTabu)
            if hasPositionFlag == False:
                break
            #得到一个新的序列
            pi1 = insertion(deepcopy(pi), i, n)
            curRes, curOutTime, curCutTime = RSP(pi1, CutLineOfPlate, StackOfPlate, L, N, S, DT, CT, OT, ES, P, B, RT)
            #如果新的序列更好，更新，并且退出
            if curRes <= bestRes:
                pi = pi1
                bestRes = curRes
                outboundTime = curOutTime
                cutTime = curCutTime
                break
            #否则把当前位置加入禁忌表中
            else: nTabu.append(n)
        plateTabu.append(i)
        bigOmega = getDelayPlate(pi, outboundTime, cutTime, plateTabu)
    return pi, bestRes
def shake(pi, k):
    def reInsertation(pi):
        for time in range(2):#通过计算 0.05%N 的结果是最好的，所以这里是2
            #确定移动的钢板及其位置
            movedPlate = choice(pi)
            movedIndex = pi.index(movedPlate)
            #确定移动范围（前）[formerIndex,backIndex]闭区间
            if movedIndex == 0: formerIndex = movedIndex
            else:
                for formerIndex in range(movedIndex)[::-1]:
                    if StackOfPlate[pi[formerIndex]] == StackOfPlate[movedPlate]:
                        formerIndex += 1
                        break
            #确定移动范围（后）
            if movedIndex == N - 1: backIndex = movedIndex
            else:
                for backIndex in range(movedIndex + 1, N):
                    if StackOfPlate[pi[backIndex]] == StackOfPlate[movedPlate]:
                        backIndex -= 1
                        break

            #重新将钢板插入到这个范围中的任意一个位置：
            reInsertIndex = choice([i for i in range(formerIndex, backIndex + 1)])

            pi.insert(reInsertIndex, pi.pop(movedIndex))
        return pi
    def segmentReconstruction(pi):
        length = 4#规定区间值为10
        formerIndex = choice(range(N - length + 1))
        backIndex = choice(range(formerIndex, formerIndex + length))
        segment = pi[formerIndex:backIndex + 1]
        #将segment的内容按照堆位分组,形成tempPlateInStack
        tempPlateInStack = deepcopy(PlatesInStack)
        #删掉tempPlateInStack中不需要重新调的任务
        for s in PlatesInStack.keys():
            for plate in PlatesInStack[s]:
                if plate not in segment: tempPlateInStack[s].remove(plate)
        #删掉tempPlateInStack中空堆位
        for s in range(S):
            if s in tempPlateInStack.keys() and len(tempPlateInStack[s]) == 0: del tempPlateInStack[s]
        #每次从tempPlateInStack中抓一个，塞回pi
        for t in range(len(segment)):
            randomS= choice(list(tempPlateInStack.keys()))
            reInsertPlate = tempPlateInStack[randomS][-1]
            pi.remove(reInsertPlate)
            pi.insert(formerIndex + t, reInsertPlate)
            tempPlateInStack[randomS].remove(reInsertPlate)
            if len(tempPlateInStack[randomS]) == 0: del tempPlateInStack[randomS]
        return pi
    if k == 1:
        return reInsertation(pi)
    elif k == 2:
        return segmentReconstruction(pi)
def BVNS(initialChoice, tMax):
    startTime = time.time()
    #初始解
    if initialChoice == 1: pi = randomInitialSolution()
    elif initialChoice == 2: pi = stackBasedInitialSolution()
    else: pi = noDelayInitialSolution()
    #local search
    pi, bestRes = localSearch(pi)
    #print('time = ', time.time() -startTime)
    while time.time() - startTime <= tMax:
        #print(iter)
        k = 1
        #print('StartRound, bestRes = ', bestRes)
        while k <= 2:
            #print('startIter, k = %d'%k)
            sTime = time.time()
            pi1, chalRes = localSearch(shake(deepcopy(pi), k))
            # print('time = ', time.time() - sTime)
            # print('chalRes = ', chalRes)
            # print('bestRes = ', bestRes)
            if chalRes < bestRes:
                #print('updata pi')
                pi = pi1
                #print(pi,'\n')
                bestRes = chalRes
                k = 1
            else: k += 1
    return bestRes, pi
def stackBasedHeuristic():
    sTime = time.time()
    pi = stackBasedInitialSolution()
    return simplifiedRSP(pi, CutLineOfPlate, StackOfPlate, L, N, S, DT, CT, OT, ES, P, B, RT), time.time() - sTime
def noDelayHeuristic():
    sTime = time.time()
    pi = noDelayInitialSolution()
    return simplifiedRSP(pi, CutLineOfPlate, StackOfPlate, L, N, S, DT, CT, OT, ES, P, B, RT), time.time() - sTime
def resultAnalysis(pi, outboundTime, cutTime):
    global CutLineOfPlate, L, DT, N
    #确定切割线和堆场作业时间的差值
    maxOutboundTime = max(outboundTime.values())
    maxCutTime = max(cutTime.values())
    #确定bottleneck的delayTime
    for plate in pi[::-1]:
        if cutTime[plate] == maxCutTime:
            bottleneckLine = CutLineOfPlate[plate]
            break
    bottleneckFormerCutTime = 0
    bottleneckDelaySum = 0
    for plate in pi:
        if CutLineOfPlate[plate] == bottleneckLine:
            bottleneckDelaySum += max(outboundTime[plate] + DT - bottleneckFormerCutTime, 0)
            bottleneckFormerCutTime = cutTime[plate]
    #确定所有任务的delayTime
    formerCutTime = [0 for l in range(L)]
    delaySum = 0
    for plate in pi:
        delaySum += max(outboundTime[plate] + DT - formerCutTime[CutLineOfPlate[plate]], 0)
        formerCutTime[CutLineOfPlate[plate]] = cutTime[plate]


    return maxCutTime - maxOutboundTime, bottleneckDelaySum, delaySum/N
if __name__ == '__main__':
    dataPath = './data/zhaoshang real data/randomCT/'
    storage = 'Hybrid algorithm result-2.0-b7.xlsx'
    reString = 'alpha_2.0_ins_21_Job_36_b_7'
    if os.path.exists(dataPath):
        files = os.listdir(dataPath)

    for eachfile in files:
        if re.search(reString, eachfile)!= None:
            print('\n\n***********************************\n', eachfile, '\n________________________________')
            originalData(dataPath + eachfile)
            # lb
            lowPlate = []
            for s in range(S):
                if len(PlatesInStack[s]) > 0: lowPlate.append(PlatesInStack[s][0])
            lb1 = sum([MOT[i] for i in MOT.keys()]) + min([CT[i] for i in lowPlate])


            CTPerLine = [0 for line in range(L)]
            MinOTPerLine = [10000 for line in range(L)]
            for job in range(N):
                CTPerLine[CutLineOfPlate[job]] += CT[job]
                if MOT[job] < MinOTPerLine[CutLineOfPlate[job]]: MinOTPerLine[CutLineOfPlate[job]] = MOT[job]
            lb2 = max([CTPerLine[l] + MinOTPerLine[l] + DT for l in range(L)])
            lb = max(lb1, lb2)
            print('lb1=',lb1,'lb2=',lb2,'\n________________________________')
            #每条切割总工时的标准差
            stdDivCutLine = np.std(CTPerLine)

            #hybrid algorithm
            while True:
                #启发式
                stack_heu_res, stack_heu_time = stackBasedHeuristic()
                noDe_heu_res, noDe_heu_time = noDelayHeuristic()
                print('_______________________________________','\nstack_heu_res = ', stack_heu_res,'\nnoDe_heu_res =', noDe_heu_res,'\n_______________________________________')

                #BVNS
                print('First BVNS')
                makespan1, pi1 = BVNS(1, 600)
                _, outboundTime1, cutTime1 = RSP(pi1, CutLineOfPlate, StackOfPlate, L, N, S, DT, CT, OT, ES, P, B, RT)
                cutStackDiff1, bottleNeckLineWaitingTime1, averagePlateDelay1 = resultAnalysis(pi1, outboundTime1,
                                                                                               cutTime1)
                print('makespan is', makespan1, '\nCut finish after stack', cutStackDiff1,
                      'mins\nBottleNeck line wait for', bottleNeckLineWaitingTime1, 'mins\nAverage waiting time is',
                      averagePlateDelay1,'\n________________________________')

                print('Second BVNS')
                makespan2, pi2 = BVNS(2, 600)
                _, outboundTime2, cutTime2 = RSP(pi2, CutLineOfPlate, StackOfPlate, L, N, S, DT, CT, OT, ES, P, B, RT)
                cutStackDiff2, bottleNeckLineWaitingTime2, averagePlateDelay2 = resultAnalysis(pi2, outboundTime2,
                                                                                               cutTime2)
                print('makespan is', makespan2, '\nCut finish after stack', cutStackDiff2,
                      'mins\nBottleNeck line wait for', bottleNeckLineWaitingTime2, 'mins\nAverage waiting time is',
                      averagePlateDelay2,'\n________________________________')

                print('Third BVNS:')
                makespan3, pi3 = BVNS(3, 600)
                _, outboundTime3, cutTime3 = RSP(pi3, CutLineOfPlate, StackOfPlate, L, N, S, DT, CT, OT, ES, P, B, RT)
                cutStackDiff3, bottleNeckLineWaitingTime3, averagePlateDelay3 = resultAnalysis(pi3, outboundTime3,
                                                                                               cutTime3)
                print('makespan is', makespan3, '\nCut finish after stack', cutStackDiff3,
                      'mins\nBottleNeck line wait for', bottleNeckLineWaitingTime3, 'mins\nAverage waiting time is',
                      averagePlateDelay3,'\n***********************************')

                wb = load_workbook(dataPath + storage)
                sheet = wb.active
                nrow = sheet.max_row + 1
                sheet.cell(row=nrow, column=1, value=eachfile)
                sheet.cell(row=nrow, column=2, value=lb1)
                sheet.cell(row=nrow, column=3, value=lb2)
                sheet.cell(row=nrow, column=4, value=lb)

                sheet.cell(row=nrow, column=5, value=makespan1)
                sheet.cell(row=nrow, column=6, value=cutStackDiff1)
                sheet.cell(row=nrow, column=7, value=bottleNeckLineWaitingTime1)
                sheet.cell(row=nrow, column=8, value=averagePlateDelay1)

                sheet.cell(row=nrow, column=9, value=makespan2)
                sheet.cell(row=nrow, column=10, value=cutStackDiff2)
                sheet.cell(row=nrow, column=11, value=bottleNeckLineWaitingTime2)
                sheet.cell(row=nrow, column=12, value=averagePlateDelay2)

                sheet.cell(row=nrow, column=13, value=makespan3)
                sheet.cell(row=nrow, column=14, value=cutStackDiff3)
                sheet.cell(row=nrow, column=15, value=bottleNeckLineWaitingTime3)
                sheet.cell(row=nrow, column=16, value=averagePlateDelay3)

                sheet.cell(row=nrow, column=17, value=stack_heu_res)
                sheet.cell(row=nrow, column=18, value=stack_heu_time)
                sheet.cell(row=nrow, column=19, value=noDe_heu_res)
                sheet.cell(row=nrow, column=20, value=noDe_heu_time)

                sheet.cell(row=nrow, column=21, value=stdDivCutLine)

                os.remove(dataPath + storage)
                wb.save(dataPath + storage)


                del S, N, L, DT, CT, PlatesInStack, B, MOT, RT, OT, P, CutLineOfPlate, StackOfPlate, H, K, O, ES
                del lb, lb1, lb2, stdDivCutLine
                del makespan1, pi1, outboundTime1, cutTime1, cutStackDiff1, bottleNeckLineWaitingTime1, averagePlateDelay1
                del makespan2, pi2, outboundTime2, cutTime2, cutStackDiff2, bottleNeckLineWaitingTime2, averagePlateDelay2
                del makespan3, pi3, outboundTime3, cutTime3, cutStackDiff3, bottleNeckLineWaitingTime3, averagePlateDelay3
                del noDe_heu_res, noDe_heu_time, stack_heu_res, stack_heu_time
                del wb, sheet, nrow
                gc.collect()
                objgraph.show_most_common_types()
                break



